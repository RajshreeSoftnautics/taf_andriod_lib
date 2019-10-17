###############################################################
# Library to generate xlsx report
###############################################################

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side


class reportgenerator():

    def __init__(self):
        self.row = 1
        self.col = 1
        self.border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def _headingCreation(self, sheet, row, col, titleName, endCol):
        """
        Method to create xls heading with colour, bold, border style
        Args:
        sheet: active sheet
        row: row number
        col: column number
        titleName: heading name
        endCol: end column number
        """
        sheet.merge_cells(start_row=row, start_column=col,
                          end_row=row, end_column=endCol)

        currentCell = sheet.cell(row, col)
        currentCell.alignment = Alignment(horizontal='center')

        sheet.cell(row=row, column=col).fill = PatternFill(
             bgColor="DCDCDC", fill_type="lightUp")

        sheet.cell(row, col).value = titleName
        sheet.cell(row, col).font = Font(bold=True)

        sheet.cell(row, col).border = self.border
        sheet.cell(row, endCol).border = self.border

    def _dataCreation(self, sheet, dataList, row, endCol):
        """
        Method to write data in xls with bold and border.
        Args:
        sheet: active sheet
        dataList: list of tuple of data
        row: row number for style
        endCol: end column number for style
        """
        for data in dataList:
            sheet.append(data)
            for i in range(endCol):
                sheet.cell(row+1, i+1).font = Font(bold=True)
            for i in range(endCol+1):
                sheet.cell(row+1, i+1).border = self.border
            row += 1
        return row

    def xlsReport(self, commonData, suiteData, testData, fileName,
                  isMobile=False):
        """
        Method to generate xlsx report.
        Args:
        commonData: common environment details
        suiteData: suite result details
        testData: test result details
        fileName: file name
        isMobile: Test Runs with mobile(android/iOS) or not
        """
        try:
            wb = Workbook()
            summarySheet = wb.worksheets[0]
            summarySheet.title = 'Summary'

            self._headingCreation(summarySheet, self.row, self.col,
                                  "Environment Details", self.col+1)

            summarySheet.column_dimensions['A'].width = 20
            summarySheet.column_dimensions['B'].width = 20

            commonList = []
            for k, v in commonData.items():
                commonList.append((k.capitalize(), v))

            self.row = self._dataCreation(summarySheet, commonList, self.row,
                                          self.col)
            self.row = self.row + 3

            self._headingCreation(summarySheet, self.row, self.col,
                                  "Test Result Summary", self.col+2)
            resultList = []
            resultList.append(("Test Suite", "TOTAL", suiteData['total']))
            resultList.append((None, "PASS", suiteData['pass']))
            resultList.append((None, "FAIL", suiteData['fail']))

            del suiteData['total']
            del suiteData['pass']
            del suiteData['fail']

            self.row = self._dataCreation(summarySheet, resultList, self.row,
                                          self.col+1)

            summarySheet.cell(self.row+1, self.col).border = self.border

            testList = []
            total = 0
            pas = 0
            fail = 0
            for k, v in testData.items():
                total += testData[k]['total']
                pas += testData[k]['pass']
                fail += testData[k]['fail']
            testList.append(("Test Case", "TOTAL", str(total)))
            testList.append((None, "PASS", str(pas)))
            testList.append((None, "FAIL", str(fail)))

            self.row = self._dataCreation(summarySheet, testList, self.row,
                                          self.col+1)
            summarySheet.cell(self.row+1, self.col+1).font = Font(bold=True)
            summarySheet.cell(self.row+1, self.col).border = self.border
            summarySheet.cell(self.row+1, self.col+1).border = self.border
            summarySheet.cell(self.row+1, self.col+2).border = self.border

            index = 1
            self.row = 1
            self.col = 1
            for k, v in testData.items():
                if isMobile is False:
                    sheetName = k.split("-")[1]
                else:
                    sheetName = k.split("_")[0].split("-", 1)[1]
                sName = "ws" + str(index)
                sName = wb.create_sheet()
                sName = wb.worksheets[index]
                sName.title = sheetName

                colList = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
                for col in colList:
                    sName.column_dimensions[col].width = 25

                self._headingCreation(sName, self.row, self.col,
                                      "Test Result Summary", self.col+2)
                testList = []
                total = 0
                pas = 0
                fail = 0
                total += v['total']
                pas += v['pass']
                fail += v['fail']
                testList.append(("Test Case", "TOTAL", str(total)))
                testList.append((None, "PASS", str(pas)))
                testList.append((None, "FAIL", str(fail)))

                del v['total']
                del v['pass']
                del v['fail']

                self.row = self._dataCreation(sName, testList, self.row,
                                              self.col+1)
                self.row = self.row + 3

                cellNameList = ["Name", "Documentation", "Tags", "Status",
                                "Fail Message", "Start Time", "End Time",
                                "Run Time (min)"]

                for name in cellNameList:
                    self._headingCreation(sName, self.row, self.col,
                                          name, self.col)
                    self.col += 1

                testResult = []
                for key, value in v.items():
                    result = [key]
                    result.extend([(d) for d in value.values()])
                    testResult.append(tuple(result))

                self.col = 1
                self.row = self._dataCreation(sName, testResult, self.row, 0)
                for l in range(len(testResult)):
                    for i in range(8):
                        sName.cell(self.row, self.col+i).border = self.border
                    if sName.cell(self.row, 4).value == "PASS":
                        sName.cell(row=self.row, column=4).fill = PatternFill(
                             bgColor="0000FF00", fill_type="lightUp")
                    elif sName.cell(self.row, 4).value == "FAIL":
                        sName.cell(row=self.row, column=4).fill = PatternFill(
                             bgColor="00FF0000", fill_type="lightUp")

                    self.row -= 1

                self.row = self.row + len(testResult)
                suiteResult = []
                for key, value in suiteData.items():
                    if key.lower() == sheetName:
                        result = [key]
                        result.extend([(d) for d in value.values()])
                        suiteResult.append(tuple(result))
                        break
                    else:
                        continue

                self.col = 1
                self.row = self._dataCreation(sName, suiteResult, self.row, 0)

                for l in range(len(suiteResult)):
                    for i in range(8):
                        sName.cell(self.row, self.col+i).border = self.border
                    if sName.cell(self.row, 4).value == "PASS":
                        sName.cell(row=self.row, column=4).fill = PatternFill(
                             bgColor="0000FF00", fill_type="lightUp")
                    elif sName.cell(self.row, 4).value == "FAIL":
                        sName.cell(row=self.row, column=4).fill = PatternFill(
                             bgColor="00FF0000", fill_type="lightUp")

                    self.row -= 1

                self.row = self.row + len(suiteResult)

                sName.insert_rows(self.row, 2)

                index += 1
                self.row = 1
                self.col = 1
            wb.save(filename=fileName)

        except Exception as error:
            return (False, error)

        return True
